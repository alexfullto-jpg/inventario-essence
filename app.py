import json
import io
import os
import secrets
from datetime import datetime, date

from flask import Flask, request, jsonify, render_template, send_file, Response

import db

app = Flask(__name__)

# Cuando se ejecuta en la nube, la aplicación exige una contraseña antes de
# mostrar cualquier dato. Railway define RAILWAY_ENVIRONMENT automáticamente;
# APP_DATA_DIR también activa este modo para evitar publicar datos por error.
IS_CLOUD = bool(os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("APP_DATA_DIR"))
APP_USERNAME = os.environ.get("APP_USERNAME", "admin")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")


@app.before_request
def require_login():
    if not IS_CLOUD:
        return None
    if not APP_PASSWORD:
        return Response(
            "La aplicación está protegida, pero falta configurar APP_PASSWORD en Railway.",
            503,
        )
    auth = request.authorization
    if auth and secrets.compare_digest(auth.username or "", APP_USERNAME) and \
            secrets.compare_digest(auth.password or "", APP_PASSWORD):
        return None
    return Response(
        "Acceso privado.", 401,
        {"WWW-Authenticate": 'Basic realm="Inventario Essence", charset="UTF-8"'},
    )


def initialize_application():
    """Creates/migrates the database on the persistent data volume."""
    db.init_db()
    db.auto_backup()


initialize_application()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def today_str():
    return date.today().isoformat()


def recipe_for(size_ml, frag_percent):
    frag_gr = size_ml * (frag_percent / 100.0)
    alc_gr = size_ml * ((100 - frag_percent) / 100.0)
    return frag_gr, alc_gr


def frasco_key(size_ml):
    return f"frasco_f{size_ml}"


def row_to_dict(row):
    return {k: row[k] for k in row.keys()}


def get_config(conn):
    row = conn.execute("SELECT * FROM config WHERE id = 1").fetchone()
    d = row_to_dict(row)
    d["cuentas"] = json.loads(d.pop("cuentas_json") or "[]")
    saldos_iniciales = json.loads(d.pop("saldos_iniciales_json") or "{}")
    # Asegura que cada cuenta configurada tenga una entrada (0 por defecto)
    # aunque todavía no se le haya definido un saldo inicial.
    for c in d["cuentas"]:
        saldos_iniciales.setdefault(c, 0)
    d["saldosIniciales"] = saldos_iniciales
    return d


def get_fragancia(conn, codigo):
    row = conn.execute("SELECT * FROM fragancias WHERE codigo = ?", (str(codigo),)).fetchone()
    return row_to_dict(row) if row else None


def saldo_factura(total, abonos):
    abonado = sum(a["monto"] for a in abonos)
    return max(0.0, round((total - abonado) * 100) / 100)


def estado_factura(total, abonos):
    saldo = saldo_factura(total, abonos)
    if saldo <= 0:
        return "Pagada"
    if len(abonos) == 0:
        return "Pendiente"
    return "Abono parcial"


def error_response(msg, code=400):
    return jsonify({"ok": False, "error": msg}), code


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Bootstrap: everything the frontend needs to render, in one call
# ---------------------------------------------------------------------------

@app.route("/api/bootstrap")
def api_bootstrap():
    conn = db.get_conn()
    try:
        config = get_config(conn)
        fragancias = [row_to_dict(r) for r in conn.execute("SELECT * FROM fragancias ORDER BY nombre")]
        otros = [row_to_dict(r) for r in conn.execute("SELECT * FROM otros_productos ORDER BY nombre")]
        clientes = [row_to_dict(r) for r in conn.execute("SELECT * FROM clientes ORDER BY nombre")]
        ventas = [row_to_dict(r) for r in conn.execute("SELECT * FROM ventas ORDER BY fecha DESC, created_at DESC")]
        compras = [row_to_dict(r) for r in conn.execute("SELECT * FROM compras ORDER BY fecha DESC, created_at DESC")]
        gastos = [row_to_dict(r) for r in conn.execute("SELECT * FROM gastos ORDER BY fecha DESC, created_at DESC")]

        facturas_rows = [row_to_dict(r) for r in conn.execute("SELECT * FROM facturas ORDER BY fecha DESC")]
        abonos_rows = [row_to_dict(r) for r in conn.execute("SELECT * FROM abonos ORDER BY fecha DESC")]
        abonos_by_factura = {}
        for a in abonos_rows:
            abonos_by_factura.setdefault(a["factura_id"], []).append(a)
        facturas = []
        for f in facturas_rows:
            abonos = abonos_by_factura.get(f["id"], [])
            f["abonos"] = abonos
            f["saldo"] = saldo_factura(f["total"], abonos)
            f["estado"] = estado_factura(f["total"], abonos)
            facturas.append(f)

        return jsonify({
            "ok": True,
            "config": config,
            "fragancias": fragancias,
            "otrosProductos": otros,
            "clientes": clientes,
            "ventas": ventas,
            "compras": compras,
            "gastos": gastos,
            "facturas": facturas,
            "serverInfo": {"lanUrl": f"http://{db.get_lan_ip()}:5000"},
        })
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@app.route("/api/config", methods=["POST"])
def api_config_update():
    data = request.get_json(force=True) or {}
    conn = db.get_conn()
    try:
        fields = []
        values = []
        mapping = {
            "businessName": "business_name",
            "businessWhatsapp": "business_whatsapp",
            "fragPercent": "frag_percent",
            "rechargeP10": "recharge_p10",
            "rechargeP30": "recharge_p30",
            "rechargeP50": "recharge_p50",
            "alcoholStockGr": "alcohol_stock_gr",
            "alcoholCostPerGr": "alcohol_cost_per_gr",
            "frasco10Stock": "frasco_f10_stock",
            "frasco30Stock": "frasco_f30_stock",
            "frasco50Stock": "frasco_f50_stock",
            "frasco10Cost": "frasco_f10_cost",
            "frasco30Cost": "frasco_f30_cost",
            "frasco50Cost": "frasco_f50_cost",
        }
        for key, col in mapping.items():
            if key in data:
                fields.append(f"{col} = ?")
                values.append(data[key])
        if "cuentas" in data and isinstance(data["cuentas"], list) and len(data["cuentas"]) > 0:
            fields.append("cuentas_json = ?")
            values.append(json.dumps(data["cuentas"]))
        if "saldosIniciales" in data and isinstance(data["saldosIniciales"], dict):
            # Se fusiona con lo que ya había guardado, para no perder el saldo
            # inicial de una cuenta que no vino en este request puntual.
            row = conn.execute("SELECT saldos_iniciales_json FROM config WHERE id = 1").fetchone()
            actuales = json.loads((row["saldos_iniciales_json"] if row else None) or "{}")
            for k, v in data["saldosIniciales"].items():
                try:
                    actuales[str(k)] = float(v)
                except (TypeError, ValueError):
                    pass
            fields.append("saldos_iniciales_json = ?")
            values.append(json.dumps(actuales))
        if fields:
            conn.execute(f"UPDATE config SET {', '.join(fields)} WHERE id = 1", values)
            conn.commit()
        return jsonify({"ok": True, "config": get_config(conn)})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Inventario: fragancias
# ---------------------------------------------------------------------------

@app.route("/api/fragancia/<codigo>", methods=["POST"])
def api_fragancia_update(codigo):
    data = request.get_json(force=True) or {}
    conn = db.get_conn()
    try:
        f = get_fragancia(conn, codigo)
        if not f:
            return error_response("Esa fragancia no existe.", 404)
        fields = []
        values = []
        if "stockGr" in data:
            fields.append("stock_gr = ?")
            values.append(float(data["stockGr"]))
        if "costPerGr" in data:
            fields.append("cost_per_gr = ?")
            values.append(float(data["costPerGr"]))
        if "ubicacion" in data:
            fields.append("ubicacion = ?")
            values.append(str(data["ubicacion"]))
        if fields:
            values.append(str(codigo))
            conn.execute(f"UPDATE fragancias SET {', '.join(fields)} WHERE codigo = ?", values)
            conn.commit()
        return jsonify({"ok": True, "fragancia": get_fragancia(conn, codigo)})
    finally:
        conn.close()


@app.route("/api/fragancia", methods=["PUT"])
def api_fragancia_upsert_bulk():
    """Import/update fragancias from a catalog JSON file (codigo, nombre, genero, p10, p30, p50)."""
    data = request.get_json(force=True) or {}
    items = data.get("items", [])
    conn = db.get_conn()
    nuevas, actualizadas = 0, 0
    try:
        for item in items:
            codigo = str(item.get("codigo", "")).strip()
            if not codigo:
                continue
            existing = get_fragancia(conn, codigo)
            if existing:
                conn.execute(
                    "UPDATE fragancias SET nombre=?, genero=?, p10=?, p30=?, p50=? WHERE codigo=?",
                    (item.get("nombre", existing["nombre"]), item.get("genero", existing["genero"]),
                     item.get("p10", existing["p10"]), item.get("p30", existing["p30"]),
                     item.get("p50", existing["p50"]), codigo),
                )
                actualizadas += 1
            else:
                conn.execute(
                    "INSERT INTO fragancias (codigo, nombre, genero, p10, p30, p50, stock_gr, cost_per_gr, ubicacion) "
                    "VALUES (?, ?, ?, ?, ?, ?, 0, 0, '')",
                    (codigo, item.get("nombre", ""), item.get("genero", "Unisex"),
                     item.get("p10", 0), item.get("p30", 0), item.get("p50", 0)),
                )
                nuevas += 1
        conn.commit()
        return jsonify({"ok": True, "nuevas": nuevas, "actualizadas": actualizadas})
    finally:
        conn.close()


def _normalize_header(s):
    import unicodedata
    s = str(s or "")
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")  # strip accents
    return s.strip().lower()


def _find_col(headers, candidates):
    for idx, h in enumerate(headers):
        if _normalize_header(h) in candidates:
            return idx
    return None


@app.route("/api/fragancia/import-excel", methods=["POST"])
def api_fragancia_import_excel():
    """Bulk-updates stock_gr, cost_per_gr and ubicacion from a spreadsheet the
    person already has (matched by Código). Empty cells are left untouched,
    any value present replaces the current one, as requested."""
    file = request.files.get("file")
    if not file or not file.filename:
        return error_response("No se recibió ningún archivo.")

    filename = file.filename.lower()
    rows = []  # list of dicts keyed by normalized header name -> raw cell value
    try:
        if filename.endswith(".csv"):
            import csv
            text = file.read().decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            raw_rows = list(reader)
            if not raw_rows:
                return error_response("El archivo no tiene filas con datos.")
            headers = raw_rows[0]
            for r in raw_rows[1:]:
                if not any(str(c).strip() for c in r):
                    continue
                rows.append(dict(zip(headers, r)))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            raw_rows = list(ws.iter_rows(values_only=True))
            if not raw_rows:
                return error_response("El archivo no tiene filas con datos.")
            headers = raw_rows[0]
            for r in raw_rows[1:]:
                if all(c is None or str(c).strip() == "" for c in r):
                    continue
                rows.append(dict(zip(headers, r)))
    except Exception as e:
        return error_response(f"No se pudo leer ese archivo. Asegúrate de que sea un Excel (.xlsx) o CSV válido. ({e})")

    if not rows:
        return error_response("El archivo no tiene filas con datos.")

    headers = list(rows[0].keys())
    codigo_idx = _find_col(headers, {"codigo", "código", "referencia", "ref", "cod"})
    if codigo_idx is None:
        return error_response("No encontré una columna de Código/Referencia en el archivo. Revisa el encabezado de esa columna.")
    codigo_key = headers[codigo_idx]
    gramos_key = headers[_find_col(headers, {"gramos", "stock", "stock (g)", "stock_g", "gr", "g"})] \
        if _find_col(headers, {"gramos", "stock", "stock (g)", "stock_g", "gr", "g"}) is not None else None
    costo_key = headers[_find_col(headers, {"costo por gramo", "costo/g", "costo_g", "costo", "costo por g", "costxg"})] \
        if _find_col(headers, {"costo por gramo", "costo/g", "costo_g", "costo", "costo por g", "costxg"}) is not None else None
    ubicacion_key = headers[_find_col(headers, {"ubicacion", "ubicación", "lugar", "estante"})] \
        if _find_col(headers, {"ubicacion", "ubicación", "lugar", "estante"}) is not None else None

    conn = db.get_conn()
    actualizadas = 0
    no_encontradas = []
    try:
        for row in rows:
            codigo = str(row.get(codigo_key) or "").strip()
            if not codigo:
                continue
            f = get_fragancia(conn, codigo)
            if not f:
                no_encontradas.append(codigo)
                continue
            fields, values = [], []
            if gramos_key is not None:
                val = row.get(gramos_key)
                if val is not None and str(val).strip() != "":
                    try:
                        fields.append("stock_gr = ?")
                        values.append(float(val))
                    except (TypeError, ValueError):
                        pass
            if costo_key is not None:
                val = row.get(costo_key)
                if val is not None and str(val).strip() != "":
                    try:
                        fields.append("cost_per_gr = ?")
                        values.append(float(val))
                    except (TypeError, ValueError):
                        pass
            if ubicacion_key is not None:
                val = row.get(ubicacion_key)
                if val is not None and str(val).strip() != "":
                    fields.append("ubicacion = ?")
                    values.append(str(val).strip())
            if fields:
                values.append(codigo)
                conn.execute(f"UPDATE fragancias SET {', '.join(fields)} WHERE codigo = ?", values)
                actualizadas += 1
        conn.commit()
        return jsonify({"ok": True, "actualizadas": actualizadas, "noEncontradas": no_encontradas})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo importar el archivo: {e}", 500)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Inventario: otros productos (bolsas, etiquetas, etc.)
# ---------------------------------------------------------------------------

@app.route("/api/otros", methods=["POST"])
def api_otros_create():
    data = request.get_json(force=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return error_response("El nombre del producto es obligatorio.")
    unidad = (data.get("unidad") or "unidad").strip() or "unidad"
    stock = float(data.get("stock") or 0)
    costo = float(data.get("costoUnit") or 0)
    item_id = db.new_id()
    conn = db.get_conn()
    try:
        conn.execute(
            "INSERT INTO otros_productos (id, nombre, unidad, stock, costo_unit) VALUES (?, ?, ?, ?, ?)",
            (item_id, nombre, unidad, stock, costo),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM otros_productos WHERE id = ?", (item_id,)).fetchone()
        return jsonify({"ok": True, "item": row_to_dict(row)})
    finally:
        conn.close()


@app.route("/api/otros/<item_id>", methods=["POST"])
def api_otros_update(item_id):
    data = request.get_json(force=True) or {}
    conn = db.get_conn()
    try:
        row = conn.execute("SELECT * FROM otros_productos WHERE id = ?", (item_id,)).fetchone()
        if not row:
            return error_response("Ese producto no existe.", 404)
        fields, values = [], []
        if "stock" in data:
            fields.append("stock = ?")
            values.append(float(data["stock"]))
        if "costoUnit" in data:
            fields.append("costo_unit = ?")
            values.append(float(data["costoUnit"]))
        if fields:
            values.append(item_id)
            conn.execute(f"UPDATE otros_productos SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
        row = conn.execute("SELECT * FROM otros_productos WHERE id = ?", (item_id,)).fetchone()
        return jsonify({"ok": True, "item": row_to_dict(row)})
    finally:
        conn.close()


@app.route("/api/otros/<item_id>", methods=["DELETE"])
def api_otros_delete(item_id):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM otros_productos WHERE id = ?", (item_id,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Ventas (registrar pedido completo con carrito)
# ---------------------------------------------------------------------------

@app.route("/api/venta", methods=["POST"])
def api_venta_create():
    data = request.get_json(force=True) or {}
    items = data.get("items", [])
    if not items:
        return error_response("El pedido no tiene productos.")

    fecha = data.get("fecha") or today_str()
    cliente = (data.get("cliente") or "").strip()
    cliente_tel = (data.get("clienteTel") or "").strip()
    cuenta = data.get("cuenta") or "Efectivo"
    es_fiado = bool(data.get("fiado"))

    conn = db.get_conn()
    try:
        config = get_config(conn)
        frag_percent = config["frag_percent"]

        # 1) Validate every line item and aggregate consumption
        frag_needed = {}
        alc_needed = 0.0
        frasco_needed = {"frasco_f10": 0, "frasco_f30": 0, "frasco_f50": 0}
        prepared = []
        for it in items:
            codigo = str(it.get("codigo"))
            tamano = int(it.get("tamano"))
            cantidad = int(it.get("cantidad") or 0)
            precio_unit = float(it.get("precioUnit") or 0)
            precio_original = float(it.get("precioOriginal") or precio_unit)
            descuento_pct = max(0.0, min(100.0, float(it.get("descuentoPct") or 0)))
            es_recarga = bool(it.get("esRecarga"))
            if cantidad <= 0:
                return error_response("Hay un producto en el pedido con cantidad inválida.")
            f = get_fragancia(conn, codigo)
            if not f:
                return error_response(f"La fragancia con referencia {codigo} no existe.")
            frag_gr, alc_gr = recipe_for(tamano, frag_percent)
            frag_needed[codigo] = frag_needed.get(codigo, 0) + frag_gr * cantidad
            alc_needed += alc_gr * cantidad
            if not es_recarga:
                frasco_needed[frasco_key(tamano)] += cantidad
            prepared.append({
                "fragancia": f, "tamano": tamano, "cantidad": cantidad,
                "precio_unit": precio_unit, "total": precio_unit * cantidad,
                "precio_original": precio_original, "descuento_pct": descuento_pct,
                "es_recarga": es_recarga, "frag_gr": frag_gr * cantidad, "alc_gr": alc_gr * cantidad,
            })

        for codigo, needed in frag_needed.items():
            f = get_fragancia(conn, codigo)
            if f["stock_gr"] < needed - 1e-9:
                return error_response(f"No hay suficiente stock de {f['nombre']} para completar el pedido.")
        if config["alcohol_stock_gr"] < alc_needed - 1e-9:
            return error_response("No hay suficiente alcohol para completar el pedido.")
        for fk, needed in frasco_needed.items():
            if needed > 0 and config[f"{fk}_stock"] < needed:
                talla = fk.replace("frasco_f", "")
                return error_response(f"No hay suficientes frascos de {talla}ML para completar el pedido.")

        # 2) Commit: deduct inventory and insert one venta row per line item, sharing the folio
        folio = config["next_folio"]
        nuevas_ventas = []
        for p in prepared:
            f = p["fragancia"]
            new_frag_stock = round((f["stock_gr"] - p["frag_gr"]) * 1000) / 1000
            conn.execute("UPDATE fragancias SET stock_gr = ? WHERE codigo = ?", (new_frag_stock, f["codigo"]))

            venta_id = db.new_id()
            conn.execute(
                "INSERT INTO ventas (id, folio, fecha, codigo, nombre, tamano, cantidad, precio_unit, total, "
                "frag_used_gr, alcohol_used_gr, cliente, cliente_tel, es_recarga, frag_cost_per_gr, "
                "alcohol_cost_per_gr, frasco_cost, precio_original, descuento_pct) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (venta_id, folio, fecha, f["codigo"], f["nombre"], p["tamano"], p["cantidad"],
                 p["precio_unit"], p["total"], p["frag_gr"], p["alc_gr"], cliente, cliente_tel,
                 1 if p["es_recarga"] else 0, f["cost_per_gr"], config["alcohol_cost_per_gr"],
                 0 if p["es_recarga"] else config[f"{frasco_key(p['tamano'])}_cost"],
                 p["precio_original"], p["descuento_pct"]),
            )
            if not p["es_recarga"]:
                fk = frasco_key(p["tamano"])
                conn.execute(f"UPDATE config SET {fk}_stock = {fk}_stock - ? WHERE id = 1", (p["cantidad"],))
            nuevas_ventas.append({
                "id": venta_id, "folio": folio, "fecha": fecha, "codigo": f["codigo"], "nombre": f["nombre"],
                "tamano": p["tamano"], "cantidad": p["cantidad"], "precioUnit": p["precio_unit"],
                "total": p["total"], "esRecarga": p["es_recarga"],
                "precioOriginal": p["precio_original"], "descuentoPct": p["descuento_pct"],
            })

        conn.execute("UPDATE config SET alcohol_stock_gr = alcohol_stock_gr - ? WHERE id = 1", (alc_needed,))
        conn.execute("UPDATE config SET next_folio = next_folio + 1 WHERE id = 1")

        # 3) Factura (siempre se crea; si no es fiado queda registrada como ya pagada)
        total_pedido = sum(p["total"] for p in prepared)
        factura_id = db.new_id()
        conn.execute(
            "INSERT INTO facturas (id, folio, fecha, cliente, cliente_tel, total) VALUES (?,?,?,?,?,?)",
            (factura_id, folio, fecha, cliente, cliente_tel, total_pedido),
        )
        if not es_fiado:
            abono_id = db.new_id()
            conn.execute(
                "INSERT INTO abonos (id, factura_id, fecha, monto, cuenta, nota) VALUES (?,?,?,?,?,?)",
                (abono_id, factura_id, fecha, total_pedido, cuenta, "Pago completo al momento de la venta"),
            )

        # 4) Save/update client
        if cliente:
            existing_c = conn.execute("SELECT * FROM clientes WHERE nombre = ?", (cliente,)).fetchone()
            if existing_c:
                if cliente_tel:
                    conn.execute("UPDATE clientes SET telefono = ? WHERE nombre = ?", (cliente_tel, cliente))
            else:
                conn.execute("INSERT INTO clientes (nombre, telefono) VALUES (?, ?)", (cliente, cliente_tel))

        conn.commit()
        return jsonify({"ok": True, "folio": folio, "ventas": nuevas_ventas})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo registrar el pedido: {e}", 500)
    finally:
        conn.close()


@app.route("/api/venta/<venta_id>", methods=["DELETE"])
def api_venta_delete(venta_id):
    conn = db.get_conn()
    try:
        v = conn.execute("SELECT * FROM ventas WHERE id = ?", (venta_id,)).fetchone()
        if not v:
            return error_response("Esa venta no existe.", 404)
        v = row_to_dict(v)

        f = get_fragancia(conn, v["codigo"])
        if f:
            conn.execute("UPDATE fragancias SET stock_gr = stock_gr + ? WHERE codigo = ?",
                         (v["frag_used_gr"], v["codigo"]))
        conn.execute("UPDATE config SET alcohol_stock_gr = alcohol_stock_gr + ? WHERE id = 1",
                     (v["alcohol_used_gr"],))
        if not v["es_recarga"]:
            fk = frasco_key(v["tamano"])
            conn.execute(f"UPDATE config SET {fk}_stock = {fk}_stock + ? WHERE id = 1", (v["cantidad"],))

        conn.execute("DELETE FROM ventas WHERE id = ?", (venta_id,))

        quedan = conn.execute("SELECT COUNT(*) AS c FROM ventas WHERE folio = ?", (v["folio"],)).fetchone()["c"]
        if quedan == 0:
            conn.execute("DELETE FROM facturas WHERE folio = ?", (v["folio"],))

        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo eliminar la venta: {e}", 500)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Compras
# ---------------------------------------------------------------------------

@app.route("/api/compra", methods=["POST"])
def api_compra_create():
    data = request.get_json(force=True) or {}
    tipo = data.get("tipo")
    cantidad = float(data.get("cantidad") or 0)
    costo = float(data.get("costoTotal") or 0)
    fecha = data.get("fecha") or today_str()
    nota = (data.get("nota") or "").strip()
    cuenta = data.get("cuenta") or "Efectivo"

    if cantidad <= 0:
        return error_response("Ingresa una cantidad válida.")

    conn = db.get_conn()
    try:
        detalle, ref_codigo, ref_nombre = "", "", ""

        if tipo == "fragancia":
            codigo = str(data.get("codigo") or "")
            f = get_fragancia(conn, codigo)
            if not f:
                return error_response("Selecciona una fragancia válida.")
            new_stock = f["stock_gr"] + cantidad
            new_cost = f["cost_per_gr"]
            if costo > 0:
                prev_total = f["cost_per_gr"] * f["stock_gr"]
                new_cost = round(((prev_total + costo) / new_stock) * 100) / 100
            conn.execute("UPDATE fragancias SET stock_gr = ?, cost_per_gr = ? WHERE codigo = ?",
                         (new_stock, new_cost, codigo))
            detalle = f"+{cantidad:g} g de fragancia"
            ref_codigo, ref_nombre = f["codigo"], f["nombre"]

        elif tipo == "alcohol":
            config = get_config(conn)
            new_stock = config["alcohol_stock_gr"] + cantidad
            new_cost = config["alcohol_cost_per_gr"]
            if costo > 0:
                prev_total = config["alcohol_cost_per_gr"] * config["alcohol_stock_gr"]
                new_cost = round(((prev_total + costo) / new_stock) * 100) / 100
            conn.execute("UPDATE config SET alcohol_stock_gr = ?, alcohol_cost_per_gr = ? WHERE id = 1",
                         (new_stock, new_cost))
            detalle = f"+{cantidad:g} g de alcohol"
            ref_nombre = "Alcohol"

        elif tipo == "otro":
            item_id = data.get("otroId")
            row = conn.execute("SELECT * FROM otros_productos WHERE id = ?", (item_id,)).fetchone()
            if not row:
                return error_response("Selecciona un producto válido (agrégalo primero en Inventario).")
            item = row_to_dict(row)
            new_stock = item["stock"] + cantidad
            new_cost = item["costo_unit"]
            if costo > 0:
                prev_total = item["costo_unit"] * item["stock"]
                new_cost = round(((prev_total + costo) / new_stock) * 100) / 100
            conn.execute("UPDATE otros_productos SET stock = ?, costo_unit = ? WHERE id = ?",
                         (new_stock, new_cost, item_id))
            detalle = f"+{cantidad:g} {item['unidad']} de {item['nombre']}"
            ref_nombre = item["nombre"]

        elif tipo in ("frasco10", "frasco30", "frasco50"):
            size = tipo.replace("frasco", "")
            fk = f"frasco_f{size}"
            conn.execute(f"UPDATE config SET {fk}_stock = {fk}_stock + ? WHERE id = 1", (int(cantidad),))
            detalle = f"+{int(cantidad)} frascos de {size}ML"
            ref_nombre = f"Frascos {size}ML"
        else:
            return error_response("Tipo de compra no reconocido.")

        compra_id = db.new_id()
        conn.execute(
            "INSERT INTO compras (id, fecha, tipo, codigo, nombre, cantidad, costo_total, nota, detalle, cuenta) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (compra_id, fecha, tipo, ref_codigo, ref_nombre, cantidad, costo, nota, detalle, cuenta),
        )
        conn.commit()
        return jsonify({"ok": True, "id": compra_id, "detalle": detalle})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo registrar la compra: {e}", 500)
    finally:
        conn.close()


@app.route("/api/compra/<compra_id>", methods=["DELETE"])
def api_compra_delete(compra_id):
    conn = db.get_conn()
    try:
        c = conn.execute("SELECT * FROM compras WHERE id = ?", (compra_id,)).fetchone()
        if not c:
            return error_response("Esa compra no existe.", 404)
        c = row_to_dict(c)

        if c["tipo"] == "fragancia":
            conn.execute("UPDATE fragancias SET stock_gr = MAX(0, stock_gr - ?) WHERE codigo = ?",
                         (c["cantidad"], c["codigo"]))
        elif c["tipo"] == "alcohol":
            conn.execute("UPDATE config SET alcohol_stock_gr = MAX(0, alcohol_stock_gr - ?) WHERE id = 1",
                         (c["cantidad"],))
        elif c["tipo"] in ("frasco10", "frasco30", "frasco50"):
            size = c["tipo"].replace("frasco", "")
            fk = f"frasco_f{size}"
            conn.execute(f"UPDATE config SET {fk}_stock = MAX(0, {fk}_stock - ?) WHERE id = 1", (int(c["cantidad"]),))
        elif c["tipo"] == "otro":
            row = conn.execute("SELECT * FROM otros_productos WHERE nombre = ?", (c["nombre"],)).fetchone()
            if row:
                conn.execute("UPDATE otros_productos SET stock = MAX(0, stock - ?) WHERE id = ?",
                             (c["cantidad"], row["id"]))

        conn.execute("DELETE FROM compras WHERE id = ?", (compra_id,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo eliminar la compra: {e}", 500)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Facturas / abonos
# ---------------------------------------------------------------------------

@app.route("/api/abono", methods=["POST"])
def api_abono_create():
    data = request.get_json(force=True) or {}
    factura_id = data.get("facturaId")
    monto = float(data.get("monto") or 0)
    cuenta = data.get("cuenta") or "Efectivo"
    fecha = data.get("fecha") or today_str()
    nota = (data.get("nota") or "").strip()

    conn = db.get_conn()
    try:
        f = conn.execute("SELECT * FROM facturas WHERE id = ?", (factura_id,)).fetchone()
        if not f:
            return error_response("Esa factura no existe.", 404)
        f = row_to_dict(f)
        abonos = [row_to_dict(a) for a in conn.execute(
            "SELECT * FROM abonos WHERE factura_id = ?", (factura_id,))]
        saldo = saldo_factura(f["total"], abonos)

        if monto <= 0:
            return error_response("Ingresa un monto válido.")
        if monto > saldo + 0.01:
            return error_response(f"El abono no puede superar el saldo pendiente (${saldo:,.0f}).")

        abono_id = db.new_id()
        conn.execute(
            "INSERT INTO abonos (id, factura_id, fecha, monto, cuenta, nota) VALUES (?,?,?,?,?,?)",
            (abono_id, factura_id, fecha, monto, cuenta, nota),
        )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo registrar el abono: {e}", 500)
    finally:
        conn.close()


@app.route("/api/factura/<factura_id>", methods=["PATCH"])
def api_factura_update(factura_id):
    """Corrige el nombre/teléfono del cliente de una factura ya registrada.
    Mantiene sincronizados los registros de ventas con ese mismo folio y la
    lista de clientes, para que el historial y las estadísticas por cliente
    queden consistentes."""
    data = request.get_json(force=True) or {}
    conn = db.get_conn()
    try:
        f = conn.execute("SELECT * FROM facturas WHERE id = ?", (factura_id,)).fetchone()
        if not f:
            return error_response("Esa factura no existe.", 404)
        f = row_to_dict(f)

        cliente = data.get("cliente")
        cliente_tel = data.get("clienteTel")
        if cliente is None and cliente_tel is None:
            return error_response("No hay cambios para guardar.")

        cliente = cliente.strip() if cliente is not None else f["cliente"]
        cliente_tel = cliente_tel.strip() if cliente_tel is not None else f["cliente_tel"]

        conn.execute("UPDATE facturas SET cliente = ?, cliente_tel = ? WHERE id = ?",
                     (cliente, cliente_tel, factura_id))
        conn.execute("UPDATE ventas SET cliente = ?, cliente_tel = ? WHERE folio = ?",
                     (cliente, cliente_tel, f["folio"]))

        if cliente:
            existing_c = conn.execute("SELECT * FROM clientes WHERE nombre = ?", (cliente,)).fetchone()
            if existing_c:
                if cliente_tel:
                    conn.execute("UPDATE clientes SET telefono = ? WHERE nombre = ?", (cliente_tel, cliente))
            else:
                conn.execute("INSERT INTO clientes (nombre, telefono) VALUES (?, ?)", (cliente, cliente_tel or ""))

        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo actualizar la factura: {e}", 500)
    finally:
        conn.close()


@app.route("/api/abono/<abono_id>", methods=["PATCH"])
def api_abono_update(abono_id):
    """Corrige un pago/abono ya registrado (por ejemplo, la cuenta a la que
    realmente llegó el dinero, el monto, la fecha o la nota)."""
    data = request.get_json(force=True) or {}
    conn = db.get_conn()
    try:
        a = conn.execute("SELECT * FROM abonos WHERE id = ?", (abono_id,)).fetchone()
        if not a:
            return error_response("Ese pago no existe.", 404)
        a = row_to_dict(a)
        factura = conn.execute("SELECT * FROM facturas WHERE id = ?", (a["factura_id"],)).fetchone()
        if not factura:
            return error_response("La factura asociada a ese pago no existe.", 404)
        factura = row_to_dict(factura)

        monto = float(data["monto"]) if data.get("monto") not in (None, "") else a["monto"]
        cuenta = (data.get("cuenta") or a["cuenta"]).strip()
        fecha = data.get("fecha") or a["fecha"]
        nota = data.get("nota") if data.get("nota") is not None else a["nota"]

        if monto <= 0:
            return error_response("El monto del pago debe ser mayor a 0.")

        otros_abonos = [row_to_dict(r) for r in conn.execute(
            "SELECT * FROM abonos WHERE factura_id = ? AND id != ?", (a["factura_id"], abono_id))]
        abonado_otros = sum(x["monto"] for x in otros_abonos)
        if monto > (factura["total"] - abonado_otros) + 0.01:
            maximo = round(factura["total"] - abonado_otros)
            return error_response(f"Ese monto supera el total de la factura (máximo ${maximo:,.0f}).")

        conn.execute(
            "UPDATE abonos SET monto = ?, cuenta = ?, fecha = ?, nota = ? WHERE id = ?",
            (monto, cuenta, fecha, nota, abono_id),
        )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo actualizar el pago: {e}", 500)
    finally:
        conn.close()


@app.route("/api/abono/<abono_id>", methods=["DELETE"])
def api_abono_delete(abono_id):
    """Elimina un pago registrado por error (por ejemplo, una venta que quedó
    marcada como pagada sin haberlo estado). La factura recupera el saldo
    pendiente correspondiente."""
    conn = db.get_conn()
    try:
        a = conn.execute("SELECT * FROM abonos WHERE id = ?", (abono_id,)).fetchone()
        if not a:
            return error_response("Ese pago no existe.", 404)
        conn.execute("DELETE FROM abonos WHERE id = ?", (abono_id,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo eliminar el pago: {e}", 500)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Gastos generales del negocio (arriendo, transporte, publicidad, etc.)
# ---------------------------------------------------------------------------

@app.route("/api/gasto", methods=["POST"])
def api_gasto_create():
    data = request.get_json(force=True) or {}
    fecha = data.get("fecha") or today_str()
    categoria = (data.get("categoria") or "Otro").strip() or "Otro"
    descripcion = (data.get("descripcion") or "").strip()
    monto = float(data.get("monto") or 0)
    cuenta = data.get("cuenta") or "Efectivo"
    if monto <= 0:
        return error_response("El monto del gasto debe ser mayor a 0.")

    gasto_id = db.new_id()
    conn = db.get_conn()
    try:
        conn.execute(
            "INSERT INTO gastos (id, fecha, categoria, descripcion, monto, cuenta) VALUES (?,?,?,?,?,?)",
            (gasto_id, fecha, categoria, descripcion, monto, cuenta),
        )
        conn.commit()
        return jsonify({"ok": True, "id": gasto_id})
    except Exception as e:
        conn.rollback()
        return error_response(f"No se pudo registrar el gasto: {e}", 500)
    finally:
        conn.close()


@app.route("/api/gasto/<gasto_id>", methods=["DELETE"])
def api_gasto_delete(gasto_id):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM gastos WHERE id = ?", (gasto_id,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Backup / restore (JSON) and Excel export
# ---------------------------------------------------------------------------

@app.route("/api/backup/export")
def api_backup_export():
    conn = db.get_conn()
    try:
        tables = ["config", "fragancias", "otros_productos", "clientes", "ventas", "compras", "gastos", "facturas", "abonos"]
        dump = {}
        for t in tables:
            dump[t] = [row_to_dict(r) for r in conn.execute(f"SELECT * FROM {t}")]
        buf = io.BytesIO(json.dumps(dump, ensure_ascii=False, indent=2).encode("utf-8"))
        buf.seek(0)
        fname = f"respaldo_inventario_{today_str()}.json"
        return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/json")
    finally:
        conn.close()


@app.route("/api/backup/import", methods=["POST"])
def api_backup_import():
    payload = request.get_json(force=True) or {}
    conn = db.get_conn()
    try:
        tables = ["abonos", "facturas", "gastos", "compras", "ventas", "clientes", "otros_productos", "fragancias", "config"]
        for t in tables:
            conn.execute(f"DELETE FROM {t}")
        for t in ["config", "fragancias", "otros_productos", "clientes", "ventas", "compras", "gastos", "facturas", "abonos"]:
            rows = payload.get(t, [])
            for r in rows:
                cols = list(r.keys())
                placeholders = ",".join(["?"] * len(cols))
                col_names = ",".join(cols)
                conn.execute(f"INSERT INTO {t} ({col_names}) VALUES ({placeholders})", [r[c] for c in cols])
        if not payload.get("config"):
            conn.execute("INSERT INTO config (id) VALUES (1)")
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return error_response(f"El archivo de respaldo no es válido: {e}", 400)
    finally:
        conn.close()


@app.route("/api/export/excel")
def api_export_excel():
    from openpyxl import Workbook

    conn = db.get_conn()
    try:
        wb = Workbook()

        ws = wb.active
        ws.title = "Ventas"
        ws.append(["Fecha", "Folio", "Codigo", "Nombre", "Tamano_ML", "Cantidad", "Precio_Original",
                   "Descuento_Pct", "Precio_Unit", "Total", "Fragancia_Usada_g", "Alcohol_Usado_g", "Cliente"])
        for v in conn.execute("SELECT * FROM ventas ORDER BY fecha DESC"):
            ws.append([v["fecha"], v["folio"], v["codigo"], v["nombre"], v["tamano"], v["cantidad"],
                       v["precio_original"], v["descuento_pct"], v["precio_unit"], v["total"],
                       v["frag_used_gr"], v["alcohol_used_gr"], v["cliente"]])

        ws2 = wb.create_sheet("Compras")
        ws2.append(["Fecha", "Tipo", "Codigo", "Nombre", "Detalle", "Cantidad", "Costo_Total", "Cuenta", "Nota"])
        for c in conn.execute("SELECT * FROM compras ORDER BY fecha DESC"):
            ws2.append([c["fecha"], c["tipo"], c["codigo"], c["nombre"], c["detalle"], c["cantidad"],
                        c["costo_total"], c["cuenta"], c["nota"]])

        ws3 = wb.create_sheet("Inventario Fragancias")
        ws3.append(["Codigo", "Nombre", "Genero", "Ubicacion", "Stock_g", "Costo_g"])
        for f in conn.execute("SELECT * FROM fragancias ORDER BY nombre"):
            ws3.append([f["codigo"], f["nombre"], f["genero"], f["ubicacion"], f["stock_gr"], f["cost_per_gr"]])

        ws4 = wb.create_sheet("Otros Productos")
        ws4.append(["Nombre", "Unidad", "Stock", "Costo_Unidad", "Valor_Total"])
        for o in conn.execute("SELECT * FROM otros_productos ORDER BY nombre"):
            ws4.append([o["nombre"], o["unidad"], o["stock"], o["costo_unit"], o["stock"] * o["costo_unit"]])

        ws_gastos = wb.create_sheet("Gastos")
        ws_gastos.append(["Fecha", "Categoria", "Descripcion", "Monto", "Cuenta"])
        for g in conn.execute("SELECT * FROM gastos ORDER BY fecha DESC"):
            ws_gastos.append([g["fecha"], g["categoria"], g["descripcion"], g["monto"], g["cuenta"]])

        ws5 = wb.create_sheet("Facturas")
        ws5.append(["Folio", "Fecha", "Cliente", "Total", "Abonado", "Saldo", "Estado"])
        facturas_rows = [row_to_dict(r) for r in conn.execute("SELECT * FROM facturas ORDER BY fecha DESC")]
        for f in facturas_rows:
            abonos = [row_to_dict(a) for a in conn.execute(
                "SELECT * FROM abonos WHERE factura_id = ?", (f["id"],))]
            abonado = sum(a["monto"] for a in abonos)
            ws5.append([f["folio"], f["fecha"], f["cliente"], f["total"], abonado,
                        saldo_factura(f["total"], abonos), estado_factura(f["total"], abonos)])

        ws6 = wb.create_sheet("Abonos")
        ws6.append(["Folio", "Fecha", "Monto", "Cuenta", "Nota"])
        for a in conn.execute(
                "SELECT abonos.*, facturas.folio AS folio FROM abonos JOIN facturas ON abonos.factura_id = facturas.id ORDER BY abonos.fecha DESC"):
            ws6.append([a["folio"], a["fecha"], a["monto"], a["cuenta"], a["nota"]])

        ws7 = wb.create_sheet("Saldo por cuenta")
        ws7.append(["Cuenta", "Saldo inicial", "+ Abonos recibidos", "- Compras a proveedores", "- Gastos", "= Saldo actual"])
        config = get_config(conn)
        abonos_all = [row_to_dict(a) for a in conn.execute("SELECT * FROM abonos")]
        compras_all = [row_to_dict(c) for c in conn.execute("SELECT * FROM compras")]
        gastos_all = [row_to_dict(g) for g in conn.execute("SELECT * FROM gastos")]
        for cuenta in config["cuentas"]:
            inicial = config["saldosIniciales"].get(cuenta, 0)
            entradas = sum(a["monto"] for a in abonos_all if a["cuenta"] == cuenta)
            salidas_compras = sum(c["costo_total"] for c in compras_all if c["cuenta"] == cuenta)
            salidas_gastos = sum(g["monto"] for g in gastos_all if g["cuenta"] == cuenta)
            actual = inicial + entradas - salidas_compras - salidas_gastos
            ws7.append([cuenta, inicial, entradas, salidas_compras, salidas_gastos, actual])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"inventario_ventas_{today_str()}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                          mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        conn.close()


if __name__ == "__main__":
    import os as _os
    import signal as _signal
    pid_path = _os.path.join(db.DATA_DIR, "server.pid")
    try:
        with open(pid_path, "w") as f:
            f.write(str(_os.getpid()))
    except Exception:
        pass

    def _cleanup_and_exit(signum, frame):
        try:
            if _os.path.exists(pid_path):
                _os.remove(pid_path)
        except Exception:
            pass
        _os._exit(0)

    try:
        _signal.signal(_signal.SIGTERM, _cleanup_and_exit)
        _signal.signal(_signal.SIGINT, _cleanup_and_exit)
    except Exception:
        pass  # not all signals are available on every platform

    lan_ip = db.get_lan_ip()
    print("=" * 60)
    print(" Essence Collection - Inventario y Ventas")
    print(" En este computador:  http://127.0.0.1:5000")
    print(f" Desde tu celular (misma WiFi):  http://{lan_ip}:5000")
    print("=" * 60)
    try:
        app.run(host="0.0.0.0", port=int(_os.environ.get("PORT", "5000")), debug=False)
    finally:
        try:
            if _os.path.exists(pid_path):
                _os.remove(pid_path)
        except Exception:
            pass
